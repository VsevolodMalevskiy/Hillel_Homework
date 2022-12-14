# Прочитать сохранённый csv-файл из задания №17 и сохранить данные
# в excel-файл, кроме возраста – столбец с этими данными не нужен.

import openpyxl
import csv

with open('data.csv', "r", encoding='utf-8') as data_csv:
    file_reader = csv.reader(data_csv)

    data_massiv = []
    for num, row in enumerate(file_reader):
        row.pop(2)
        data_massiv.insert(num, row)

# book = openpyxl.load_workbook(filename='test.xlsx')
#
# sheet.insert_index(0)  # вставляет строку выше указанного индекса (в самый верх)
# book.close()

book = openpyxl.Workbook()


sheet = book.active
sheet.title = 'Persons'
sheet_2 = book.create_sheet('2')  # Создание второй страницы с названием 2
print(book.sheetnames) # Печать наименования страниц

headline = []
for item in range(6):
    headline.append(f"Person {item+1}")

sheet.append(('', *headline))

for index, row in enumerate(data_massiv):
    for index_in, column in enumerate(data_massiv[index]):
        sheet[index_in+2][index].value = column

book.save('test.xlsx')
book.close()

